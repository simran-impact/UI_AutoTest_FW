import openpyxl
import pandas

from configurations.baseConfig import *


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value


def writeData(file, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(file)


def read_uniqueValues_inACol(filePath, strColName):
    lst_uniqueValues = []
    try:
        data_retrieved = pandas.read_csv(filePath)
        lst_uniqueValues = data_retrieved[strColName].unique().tolist()
    except Exception as e:
        print(e)
    return lst_uniqueValues


def filter_using_ColValues(filepath, strColName, strColValue):
    df = pandas.read_csv(filepath)
    df_new = df[df[strColName] == strColValue]
    df_new.to_csv(EXTRACTEDDATAFILES_FLD_PATH + strColValue + '_filteredColValues.csv')
    return df_new


def read_uniqueValues_from_Col1_respectTo_Col2Value(filePath, lstRefColName, strColValue):
    lst_mapRequiredData = {}
    try:
        data_retrieved = pandas.read_csv(filePath, usecols=lstRefColName)
        lst_retrievedValues = data_retrieved.values.tolist()
        lst_mapData = {}
        for i in lst_retrievedValues:
            if i[1] in lst_mapData.keys():
                lst_mapData[i[1]].add(i[0])  # add all Col2 values  in loop
            else:
                lst_mapData.update({i[1]: {i[0]}})  # add first Col2 value to key (Col1 value)
        lst_mapRequiredData = lst_mapData.get(strColValue)
    except Exception as e:
        print(e)
    return lst_mapRequiredData
